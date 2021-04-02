import openpyxl

from utilities.search_login_data import SearchLoginData

class XlsxReader:
    #This method is to read data from Login Data sheet
    @staticmethod
    def get_xlsx_login_data():
        wb = openpyxl.load_workbook(".//TestData//LoginData.xlsx")
        sheet = wb['Sheet1']
        max_col = sheet.max_column
        data = []
        for i in range(1, max_col + 1):
            search_login_data = SearchLoginData(sheet.cell(2, i).value,  # username
                                                sheet.cell(2, i).value)  # password
            data.append(search_login_data)
            return data
