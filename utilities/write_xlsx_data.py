import openpyxl

from utilities.search_login_data import SearchLoginData

class XlsxWriter:
    #This method is to read data from Login Data sheet
    @staticmethod
    def write_xlsx_login_data():
        file = ".//TestData//LoginData.xlsx"
        wb = openpyxl.load_workbook(file)
        sheet = wb['Sheet1']
        max_col = sheet.max_column
        sheet.cell(row=2, column=3).value = "Pass"
        wb.save(file)


